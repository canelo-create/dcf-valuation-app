"""Generate dcf-inditex.xlsx — full DCF model with formulas, scenarios, sensitivity."""

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
INPUTS = json.loads((ROOT / "dcf-inputs.json").read_text(encoding="utf-8"))
XLSX_PATH = ROOT / "dcf-inditex.xlsx"

NAVY = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
LIGHTBLUE = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
LIGHTGREY = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
CENTER_BLUE = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

FONT_DATA = Font(name="Times New Roman", size=11, color="000000")
FONT_INPUT = Font(name="Times New Roman", size=11, color="0070C0")
FONT_HDR_WHITE = Font(name="Times New Roman", size=12, color="FFFFFF", bold=True)
FONT_HDR_BLACK = Font(name="Times New Roman", size=11, color="000000", bold=True)
FONT_TITLE = Font(name="Times New Roman", size=14, bold=True)
FONT_CENTER_BOLD = Font(name="Times New Roman", size=11, color="000000", bold=True)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

SOURCE_NOTE = "Source: stockanalysis.com (TTM and historical), accessed 2026-05-13. URL in case study README."


def section_header(ws, row, text, ncols):
    ws.cell(row=row, column=1, value=text)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = NAVY
        cell.font = FONT_HDR_WHITE
        cell.alignment = CENTER


def label_cell(ws, row, col, text, bold=False, fill=None):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = FONT_HDR_BLACK if bold else FONT_DATA
    cell.alignment = LEFT
    if fill:
        cell.fill = fill
    return cell


def input_cell(ws, row, col, value, fmt=None, note=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = FONT_INPUT
    cell.alignment = CENTER
    if fmt:
        cell.number_format = fmt
    if note:
        cell.comment = Comment(note, "Build script")
    return cell


def formula_cell(ws, row, col, formula, fmt=None, bold=False):
    cell = ws.cell(row=row, column=col, value=formula)
    cell.font = FONT_CENTER_BOLD if bold else FONT_DATA
    cell.alignment = CENTER
    if fmt:
        cell.number_format = fmt
    return cell


def col_header(ws, row, labels, start_col=1):
    for i, label in enumerate(labels):
        c = start_col + i
        cell = ws.cell(row=row, column=c, value=label)
        cell.fill = LIGHTBLUE
        cell.font = FONT_HDR_BLACK
        cell.alignment = CENTER
    return start_col + len(labels) - 1


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "DCF Model"

    md = INPUTS["market_data"]
    wacc_d = INPUTS["wacc"]
    base = INPUTS["projection_assumptions_base"]
    bear = INPUTS["projection_assumptions_bear"]
    bull = INPUTS["projection_assumptions_bull"]
    hist = INPUTS["historical"]

    # Row 1-3: Title block
    ws.cell(row=1, column=1, value="INDITEX (ITX.MC) - DISCOUNTED CASH FLOW VALUATION")
    ws.cell(row=1, column=1).font = FONT_TITLE
    ws.cell(row=2, column=1, value="As of 2026-05-13 | All figures in EUR Millions except per-share | Fiscal year ends January 31")
    ws.cell(row=3, column=1, value="Base case: 10-year explicit projection, WACC 8.26%, Terminal g 2.5%, Exit multiple 12.0x EV/EBITDA")

    # === SECTION 1: MARKET DATA & KEY INPUTS (rows 5-14) ===
    section_header(ws, 5, "MARKET DATA & KEY INPUTS", 12)
    inputs_data = [
        ("Current share price (EUR)", md["current_price_eur"], "0.00", SOURCE_NOTE),
        ("Shares outstanding (M)", md["shares_outstanding_m"], "#,##0", SOURCE_NOTE),
        ("Market capitalization (EUR M)", md["market_cap_eur_m"], "#,##0", "=Price * Shares"),
        ("Cash and ST investments (EUR M)", md["cash_and_st_investments_eur_m"], "#,##0", SOURCE_NOTE),
        ("Total financial debt (EUR M)", md["total_financial_debt_eur_m"], "#,##0", SOURCE_NOTE),
        ("Net cash (EUR M)", md["net_cash_eur_m"], "#,##0", "Cash minus debt. Positive means net cash."),
        ("Enterprise value (current, EUR M)", md["enterprise_value_eur_m"], "#,##0", "Market cap minus net cash"),
        ("Beta (5Y)", md["beta_5y"], "0.00", SOURCE_NOTE),
    ]
    for i, (label, val, fmt, note) in enumerate(inputs_data):
        r = 6 + i
        label_cell(ws, r, 1, label)
        input_cell(ws, r, 3, val, fmt=fmt, note=note)
    # Named cell reference rows for WACC block
    PRICE_ROW = 6
    SHARES_ROW = 7
    NETCASH_ROW = 11

    # === SECTION 2: WACC BUILD (rows 16-25) ===
    section_header(ws, 16, "WACC BUILD (CAPM)", 12)
    wacc_data = [
        ("Risk-free rate (Spain 10Y bond)", wacc_d["risk_free_rate_pct"] / 100, "0.00%", wacc_d["risk_free_source"]),
        ("Equity risk premium (Damodaran 2026 Spain)", wacc_d["equity_risk_premium_pct"] / 100, "0.00%", wacc_d["erp_source"]),
        ("Beta (5Y)", wacc_d["beta"], "0.00", "Per stockanalysis.com 5-year regression"),
    ]
    for i, (label, val, fmt, note) in enumerate(wacc_data):
        r = 17 + i
        label_cell(ws, r, 1, label)
        input_cell(ws, r, 3, val, fmt=fmt, note=note)
    RF_ROW = 17
    ERP_ROW = 18
    BETA_ROW = 19

    label_cell(ws, 20, 1, "Cost of equity = Rf + Beta * ERP", bold=True)
    formula_cell(ws, 20, 3, f"=C{RF_ROW}+C{BETA_ROW}*C{ERP_ROW}", fmt="0.00%", bold=True)

    label_cell(ws, 21, 1, "Pretax cost of debt")
    input_cell(ws, 21, 3, wacc_d["pretax_cost_of_debt_pct"] / 100, fmt="0.00%", note="Estimated, low importance due to minimal debt")

    label_cell(ws, 22, 1, "Tax rate")
    input_cell(ws, 22, 3, wacc_d["tax_rate_pct"] / 100, fmt="0.00%", note="Based on FY25 effective tax rate")
    TAX_ROW = 22

    label_cell(ws, 23, 1, "After-tax cost of debt", bold=True)
    formula_cell(ws, 23, 3, f"=C21*(1-C{TAX_ROW})", fmt="0.00%", bold=True)

    label_cell(ws, 24, 1, "WACC (effective 100% equity, net cash position)", bold=True, fill=LIGHTBLUE)
    formula_cell(ws, 24, 3, f"=C20", fmt="0.00%", bold=True).fill = LIGHTBLUE
    WACC_ROW = 24

    # === SECTION 3: HISTORICAL FINANCIALS (rows 27-36) ===
    section_header(ws, 27, "HISTORICAL FINANCIALS (5Y)", 12)
    hist_headers = ["Metric"] + [f"FY{y}" for y in ["2021", "2022", "2023", "2024", "2025"]] + ["5Y CAGR"]
    for c, label in enumerate(hist_headers):
        cell = ws.cell(row=28, column=c + 1, value=label)
        cell.fill = LIGHTBLUE
        cell.font = FONT_HDR_BLACK
        cell.alignment = CENTER
    hist_rows = [
        ("Revenue", "revenue", "#,##0"),
        ("EBITDA", "ebitda", "#,##0"),
        ("EBIT", "ebit", "#,##0"),
        ("D&A", "da", "#,##0"),
        ("CapEx", "capex", "#,##0"),
        ("FCF", "fcf", "#,##0"),
        ("Net Income", "net_income", "#,##0"),
    ]
    for i, (label, key, fmt) in enumerate(hist_rows):
        r = 29 + i
        label_cell(ws, r, 1, label, bold=True)
        for j, fy in enumerate(["FY2021", "FY2022", "FY2023", "FY2024", "FY2025"]):
            val = hist[fy].get(key, 0)
            input_cell(ws, r, 2 + j, val, fmt=fmt, note=f"{fy} from stockanalysis.com financials")
        # CAGR FY21-FY25
        formula_cell(ws, r, 7, f"=(F{r}/B{r})^(1/4)-1", fmt="0.0%")
    # Margin row
    margin_rows = [
        ("EBITDA margin", "ebitda", "revenue", "0.0%"),
        ("FCF margin", "fcf", "revenue", "0.0%"),
    ]
    for i, (label, num_key, den_key, fmt) in enumerate(margin_rows):
        r = 36 + i
        label_cell(ws, r, 1, label, bold=True, fill=LIGHTGREY)
        for j, fy in enumerate(["FY2021", "FY2022", "FY2023", "FY2024", "FY2025"]):
            num_idx = next((idx for idx, t in enumerate(hist_rows) if t[1] == num_key), 0)
            den_idx = next((idx for idx, t in enumerate(hist_rows) if t[1] == den_key), 0)
            num_row = 29 + num_idx
            den_row = 29 + den_idx
            col = get_column_letter(2 + j)
            formula_cell(ws, r, 2 + j, f"={col}{num_row}/{col}{den_row}", fmt=fmt)
            ws.cell(row=r, column=2 + j).fill = LIGHTGREY

    # === SECTION 4: REVENUE PROJECTION BASE CASE (rows 40-58) ===
    section_header(ws, 40, "PROJECTION - BASE CASE (10Y EXPLICIT)", 12)
    years = list(range(1, 11))
    proj_headers = ["Metric"] + [f"Y{y}" for y in years] + ["Sum"]
    for c, label in enumerate(proj_headers):
        cell = ws.cell(row=41, column=c + 1, value=label)
        cell.fill = LIGHTBLUE
        cell.font = FONT_HDR_BLACK
        cell.alignment = CENTER

    PROJ_FIRST_ROW = 42
    GROWTH_ROW = 42
    REV_ROW = 43
    EBMARG_ROW = 44
    EBITDA_ROW = 45
    DA_PCT_ROW = 46
    DA_ROW = 47
    EBIT_ROW = 48
    NOPAT_ROW = 49
    CAPEX_PCT_ROW = 50
    CAPEX_ROW = 51
    NWC_ROW = 52
    FCF_ROW = 53
    DISC_PERIOD_ROW = 54
    PV_FACTOR_ROW = 55
    PV_FCF_ROW = 56

    label_cell(ws, GROWTH_ROW, 1, "Revenue growth %")
    for i in range(10):
        input_cell(ws, GROWTH_ROW, 2 + i, base["revenue_growth_yoy_pct"][i] / 100, fmt="0.0%")

    label_cell(ws, REV_ROW, 1, "Revenue (EUR M)", bold=True)
    base_rev_cell_addr = "F33"  # FY2025 revenue cell
    for i in range(10):
        if i == 0:
            formula = f"={base_rev_cell_addr}*(1+B{GROWTH_ROW})"
        else:
            prev = get_column_letter(2 + i - 1)
            cur = get_column_letter(2 + i)
            formula = f"={prev}{REV_ROW}*(1+{cur}{GROWTH_ROW})"
        formula_cell(ws, REV_ROW, 2 + i, formula, fmt="#,##0", bold=True)

    label_cell(ws, EBMARG_ROW, 1, "EBITDA margin %")
    for i in range(10):
        input_cell(ws, EBMARG_ROW, 2 + i, base["ebitda_margin_pct"][i] / 100, fmt="0.0%")

    label_cell(ws, EBITDA_ROW, 1, "EBITDA", bold=True)
    for i in range(10):
        col = get_column_letter(2 + i)
        formula_cell(ws, EBITDA_ROW, 2 + i, f"={col}{REV_ROW}*{col}{EBMARG_ROW}", fmt="#,##0", bold=True)

    label_cell(ws, DA_PCT_ROW, 1, "D&A % of revenue")
    for i in range(10):
        input_cell(ws, DA_PCT_ROW, 2 + i, base["da_pct_of_revenue"][i] / 100, fmt="0.0%")

    label_cell(ws, DA_ROW, 1, "D&A")
    for i in range(10):
        col = get_column_letter(2 + i)
        formula_cell(ws, DA_ROW, 2 + i, f"={col}{REV_ROW}*{col}{DA_PCT_ROW}", fmt="#,##0")

    label_cell(ws, EBIT_ROW, 1, "EBIT (= EBITDA - D&A)")
    for i in range(10):
        col = get_column_letter(2 + i)
        formula_cell(ws, EBIT_ROW, 2 + i, f"={col}{EBITDA_ROW}-{col}{DA_ROW}", fmt="#,##0")

    label_cell(ws, NOPAT_ROW, 1, "NOPAT (= EBIT * (1-tax))")
    for i in range(10):
        col = get_column_letter(2 + i)
        formula_cell(ws, NOPAT_ROW, 2 + i, f"={col}{EBIT_ROW}*(1-$C${TAX_ROW})", fmt="#,##0")

    label_cell(ws, CAPEX_PCT_ROW, 1, "CapEx % of revenue")
    for i in range(10):
        input_cell(ws, CAPEX_PCT_ROW, 2 + i, base["capex_pct_of_revenue"][i] / 100, fmt="0.0%")

    label_cell(ws, CAPEX_ROW, 1, "CapEx")
    for i in range(10):
        col = get_column_letter(2 + i)
        formula_cell(ws, CAPEX_ROW, 2 + i, f"={col}{REV_ROW}*{col}{CAPEX_PCT_ROW}", fmt="#,##0")

    nwc_pct = base["nwc_change_pct_of_revenue_change"] / 100
    label_cell(ws, NWC_ROW, 1, f"Change in NWC ({nwc_pct*100:.1f}% of Δrev)")
    for i in range(10):
        col = get_column_letter(2 + i)
        if i == 0:
            prev_cell = base_rev_cell_addr
        else:
            prev_cell = f"{get_column_letter(2 + i - 1)}{REV_ROW}"
        formula_cell(ws, NWC_ROW, 2 + i, f"=({col}{REV_ROW}-{prev_cell})*{nwc_pct}", fmt="#,##0")

    label_cell(ws, FCF_ROW, 1, "Unlevered FCF (= NOPAT + D&A - CapEx - ΔNWC)", bold=True, fill=LIGHTBLUE)
    for i in range(10):
        col = get_column_letter(2 + i)
        formula_cell(ws, FCF_ROW, 2 + i, f"={col}{NOPAT_ROW}+{col}{DA_ROW}-{col}{CAPEX_ROW}-{col}{NWC_ROW}", fmt="#,##0", bold=True).fill = LIGHTBLUE

    label_cell(ws, DISC_PERIOD_ROW, 1, "Discount period (mid-year)")
    for i in range(10):
        formula_cell(ws, DISC_PERIOD_ROW, 2 + i, i + 0.5, fmt="0.0")

    label_cell(ws, PV_FACTOR_ROW, 1, "Discount factor")
    for i in range(10):
        col = get_column_letter(2 + i)
        formula_cell(ws, PV_FACTOR_ROW, 2 + i, f"=1/((1+$C${WACC_ROW})^{col}{DISC_PERIOD_ROW})", fmt="0.0000")

    label_cell(ws, PV_FCF_ROW, 1, "PV of FCF", bold=True, fill=LIGHTBLUE)
    for i in range(10):
        col = get_column_letter(2 + i)
        formula_cell(ws, PV_FCF_ROW, 2 + i, f"={col}{FCF_ROW}*{col}{PV_FACTOR_ROW}", fmt="#,##0", bold=True).fill = LIGHTBLUE
    # Sum of PV FCFs
    formula_cell(ws, PV_FCF_ROW, 12, f"=SUM(B{PV_FCF_ROW}:K{PV_FCF_ROW})", fmt="#,##0", bold=True).fill = LIGHTBLUE

    # === SECTION 5: TERMINAL VALUE + EQUITY BRIDGE (rows 58-72) ===
    section_header(ws, 58, "TERMINAL VALUE & EQUITY BRIDGE", 12)

    label_cell(ws, 59, 1, "Final year FCF (Y10)")
    formula_cell(ws, 59, 3, f"=K{FCF_ROW}", fmt="#,##0")

    label_cell(ws, 60, 1, "Final year EBITDA (Y10)")
    formula_cell(ws, 60, 3, f"=K{EBITDA_ROW}", fmt="#,##0")

    label_cell(ws, 61, 1, "Terminal growth rate")
    input_cell(ws, 61, 3, base["terminal_growth_pct"] / 100, fmt="0.0%", note="Base case assumption")
    TG_ROW = 61

    label_cell(ws, 62, 1, "Exit multiple (EV/EBITDA)")
    input_cell(ws, 62, 3, base["exit_multiple_ev_ebitda"], fmt='0.0"x"', note="Base case assumption")
    EM_ROW = 62

    label_cell(ws, 63, 1, "Terminal Value (perpetuity, gross)")
    formula_cell(ws, 63, 3, f"=C59*(1+C{TG_ROW})/($C${WACC_ROW}-C{TG_ROW})", fmt="#,##0")

    label_cell(ws, 64, 1, "Terminal Value (exit multiple, gross)")
    formula_cell(ws, 64, 3, f"=C60*C{EM_ROW}", fmt="#,##0")

    label_cell(ws, 65, 1, "PV of Terminal (blended avg)", bold=True)
    formula_cell(ws, 65, 3, f"=AVERAGE(C63,C64)/((1+$C${WACC_ROW})^K{DISC_PERIOD_ROW})", fmt="#,##0", bold=True)

    label_cell(ws, 67, 1, "Sum PV of FCFs (Y1-Y10)", bold=True)
    formula_cell(ws, 67, 3, f"=L{PV_FCF_ROW}", fmt="#,##0", bold=True)

    label_cell(ws, 68, 1, "Enterprise Value (= PV FCFs + PV Terminal)", bold=True, fill=LIGHTBLUE)
    formula_cell(ws, 68, 3, f"=C67+C65", fmt="#,##0", bold=True).fill = LIGHTBLUE

    label_cell(ws, 69, 1, "Net cash (add back)")
    formula_cell(ws, 69, 3, f"=C{NETCASH_ROW}", fmt="#,##0")

    label_cell(ws, 70, 1, "Equity Value", bold=True, fill=LIGHTBLUE)
    formula_cell(ws, 70, 3, f"=C68+C69", fmt="#,##0", bold=True).fill = LIGHTBLUE

    label_cell(ws, 71, 1, "Shares outstanding (M)")
    formula_cell(ws, 71, 3, f"=C{SHARES_ROW}", fmt="#,##0")

    label_cell(ws, 72, 1, "IMPLIED SHARE PRICE (EUR)", bold=True, fill=CENTER_BLUE)
    formula_cell(ws, 72, 3, f"=C70/C71", fmt="0.00", bold=True).fill = CENTER_BLUE

    label_cell(ws, 73, 1, "Current market price (EUR)")
    formula_cell(ws, 73, 3, f"=C{PRICE_ROW}", fmt="0.00")

    label_cell(ws, 74, 1, "Upside / (Downside) vs market", bold=True, fill=CENTER_BLUE)
    formula_cell(ws, 74, 3, f"=C72/C73-1", fmt="0.0%", bold=True).fill = CENTER_BLUE

    label_cell(ws, 75, 1, "Terminal value % of EV (sanity 50-70%)")
    formula_cell(ws, 75, 3, f"=C65/C68", fmt="0.0%")

    label_cell(ws, 76, 1, "Implied EV/EBITDA (current EBITDA)")
    formula_cell(ws, 76, 3, f"=C68/F30", fmt='0.00"x"')

    # === SECTION 6: SENSITIVITY (rows 79+) ===
    section_header(ws, 79, "SENSITIVITY: IMPLIED SHARE PRICE - WACC vs TERMINAL GROWTH", 12)
    wacc_range = INPUTS["sensitivity_ranges"]["wacc_pct"]
    tg_range = INPUTS["sensitivity_ranges"]["terminal_growth_pct"]
    # Header row 80: WACC values across, terminal g down
    ws.cell(row=80, column=1, value="Terminal g \\ WACC").font = FONT_HDR_BLACK
    ws.cell(row=80, column=1).fill = LIGHTBLUE
    for c, w in enumerate(wacc_range):
        cell = ws.cell(row=80, column=2 + c, value=w / 100)
        cell.fill = LIGHTBLUE
        cell.font = FONT_HDR_BLACK
        cell.number_format = "0.00%"
        cell.alignment = CENTER
    # For sensitivity, we can't do live formulas easily for full DCF recalc with different WACC.
    # Simplification: use simple growing perpetuity on terminal year FCF for sensitivity.
    # FCF Y10 already computed at K{FCF_ROW}.
    # Implied share price approx = (sum PV FCFs at base WACC + (FCF_Y10 * (1+g) / (W-g)) / (1+W)^9.5 + net cash) / shares
    # For simpler sensitivity: compute terminal value at varied W and g, hold sum PV FCFs constant approx.
    # Better: use precomputed values from compute_dcf.py for accuracy. For xlsx live model, use formula.
    for ri, tg in enumerate(tg_range):
        cell = ws.cell(row=81 + ri, column=1, value=tg / 100)
        cell.font = FONT_HDR_BLACK
        cell.fill = LIGHTBLUE
        cell.number_format = "0.00%"
        cell.alignment = CENTER
        for ci, w in enumerate(wacc_range):
            wf = w / 100
            tgf = tg / 100
            # Approx implied share: sum_pv_fcf computed at base wacc + terminal at varied W,g
            # Simplification: compute everything at the variable wacc by referencing terminal cells
            # For now use: implied_share = (FCF_Y10 * (1+g)/(W-g) / (1+W)^9.5 + sum_pv_fcf + net_cash) / shares
            # We hold sum_pv_fcf at base case (limitation noted in notes)
            formula = (
                f"=((K{FCF_ROW}*(1+{tgf}))/({wf}-{tgf})/((1+{wf})^9.5) + C67 + C{NETCASH_ROW})/C{SHARES_ROW}"
            )
            cell2 = ws.cell(row=81 + ri, column=2 + ci, value=formula)
            cell2.number_format = "0.00"
            cell2.font = FONT_DATA
            cell2.alignment = CENTER
            # Highlight base case (WACC 8.5% closest to 8.26%, tg 2.5%)
            if abs(w - 8.5) < 0.01 and abs(tg - 2.5) < 0.01:
                cell2.fill = CENTER_BLUE
                cell2.font = FONT_CENTER_BOLD

    # === SECTION 7: SCENARIO SUMMARY (rows 91-96) ===
    section_header(ws, 90, "SCENARIO SUMMARY - Bear / Base / Bull", 12)
    col_header(ws, 91, ["Scenario", "Revenue Y10", "EBITDA Margin Y10", "Terminal g", "Exit mult", "Implied Share (EUR)", "Upside vs market"], start_col=1)

    scenarios_summary = []
    # Compute scenarios in Python (already done) and inject as values for display
    import compute_dcf
    scens = compute_dcf.run_all()
    for i, case in enumerate(["bear", "base", "bull"]):
        s = scens[case]
        p = s["projection"]
        d = s["dcf"]
        tg = s["tg"]
        em = s["em"]
        row = 92 + i
        label_cell(ws, row, 1, case.upper(), bold=True)
        input_cell(ws, row, 2, p["revenue"][-1], fmt="#,##0", note="Computed externally")
        input_cell(ws, row, 3, p["ebitda"][-1] / p["revenue"][-1], fmt="0.0%", note="Computed externally")
        input_cell(ws, row, 4, tg / 100, fmt="0.0%")
        input_cell(ws, row, 5, em, fmt='0.0"x"')
        input_cell(ws, row, 6, d["implied_share_price"], fmt="0.00", note=f"{case} implied price, computed via compute_dcf.py")
        input_cell(ws, row, 7, d["implied_share_price"] / md["current_price_eur"] - 1, fmt="0.0%")
        if case == "base":
            for c in range(1, 8):
                ws.cell(row=row, column=c).fill = CENTER_BLUE

    # === SECTION 8: NOTES (rows 99+) ===
    section_header(ws, 99, "NOTES & METHODOLOGY", 12)
    notes = [
        "Methodology: 10-year explicit projection of Unlevered FCF, discounted to present at WACC. Terminal value computed two ways (perpetuity growth and exit multiple), averaged.",
        "WACC: Cost of equity via CAPM. Inditex has net cash position (EUR 11B liquid vs EUR 28M financial debt), so WACC effectively equals cost of equity = 8.26%.",
        "Risk-free rate: Spain 10Y government bond yield approx 3.2% as of May 2026. ERP 5.5% per Damodaran Spain mature market estimate.",
        "Beta 0.92 from 5Y monthly regression vs MSCI Spain. Lower than 1.0 reflects defensive consumer staple-like profile.",
        "Base case projection: revenue tapers from 7% (Y1) to 2.5% (Y10). EBITDA margin held at 28.5% (Inditex historic 28-29% range). CapEx tapers from 7% to 5% of revenue.",
        "Working capital: Inditex runs negative working capital (suppliers paid after collections). Modeled at -2% of revenue change.",
        "Terminal growth 2.5% = roughly long-run Eurozone nominal GDP estimate. Exit multiple 12x EV/EBITDA below Inditex actual 13.2x to avoid double-counting current premium.",
        "Bear case: revenue stagnation 1.5-3%, margin compression to 23%, terminal 1.5%, exit 9x. Captures global retail downturn scenario.",
        "Bull case: revenue 3.5-9% sustained, margin expansion to 30%, terminal 3%, exit 14x. Captures continued Inditex outperformance + Asia expansion.",
        "Base implied share EUR 56.27 vs current EUR 48.55 = +16% upside. Moderate buy signal.",
        "Sanity checks: Terminal % of EV = 60.8% (within 50-70% acceptable range). Implied EV/EBITDA current = 14.6x vs market 13.2x.",
        "LIMITATIONS: (a) FY2024 data extraction error from stockanalysis.com corrected manually using Inditex official press release. (b) Sensitivity table holds sum PV FCFs constant at base WACC for simplicity, only varies terminal — for full sensitivity, rerun compute_dcf.py with different WACC inputs. (c) Operating lease liabilities (IFRS 16) not separately modeled, embedded in EBITDA via D&A.",
        "DISCLAIMER (per Anthropic FSI repo README): Nothing here constitutes investment advice. Verify against Inditex official annual report and 20-F. Output is analyst work product for review.",
        "Source repo: https://github.com/anthropics/financial-services (Apache 2.0). Skill: dcf-model from financial-analysis vertical plugin.",
    ]
    for i, n in enumerate(notes):
        ws.cell(row=100 + i, column=1, value=n).font = FONT_DATA
        ws.merge_cells(start_row=100 + i, start_column=1, end_row=100 + i, end_column=12)

    # Column widths
    ws.column_dimensions["A"].width = 45
    for c in range(2, 13):
        ws.column_dimensions[get_column_letter(c)].width = 14

    wb.save(XLSX_PATH)
    print(f"Wrote {XLSX_PATH}")


if __name__ == "__main__":
    build()
