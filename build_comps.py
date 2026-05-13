"""Build comps xlsx for any target company + peer set.

Usage:
    python build_comps.py --inputs cases/inditex/inputs.json --output cases/inditex/comps.xlsx

Reads peer CSV path from inputs.json (data_sources.comps), resolves relative to inputs file.
"""

import argparse
import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

NAVY = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
LIGHTBLUE = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
LIGHTGREY = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

FONT_DATA = Font(name="Times New Roman", size=11, color="000000")
FONT_INPUT = Font(name="Times New Roman", size=11, color="0070C0")
FONT_HDR_WHITE = Font(name="Times New Roman", size=12, color="FFFFFF", bold=True)
FONT_HDR_BLACK = Font(name="Times New Roman", size=11, color="000000", bold=True)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def section_header(ws, row, text, ncols):
    ws.cell(row=row, column=1, value=text)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = NAVY
        cell.font = FONT_HDR_WHITE
        cell.alignment = CENTER


def col_header(ws, row, labels):
    for c, label in enumerate(labels, start=1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.fill = LIGHTBLUE
        cell.font = FONT_HDR_BLACK
        cell.alignment = CENTER


def write_input(ws, row, col, value, source_note=None, fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = FONT_INPUT
    cell.alignment = CENTER
    if source_note:
        cell.comment = Comment(source_note, "Build script")
    if fmt:
        cell.number_format = fmt
    return cell


def write_formula(ws, row, col, formula, fmt=None):
    cell = ws.cell(row=row, column=col, value=formula)
    cell.font = FONT_DATA
    cell.alignment = CENTER
    if fmt:
        cell.number_format = fmt
    return cell


def stats_row(ws, row, label, ranges, fmt):
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = FONT_HDR_BLACK
    cell.alignment = LEFT
    cell.fill = LIGHTGREY
    func_map = {"Max": "MAX", "75th Percentile": "QUARTILE", "Median": "MEDIAN", "25th Percentile": "QUARTILE", "Min": "MIN"}
    func = func_map[label]
    for c, rng in ranges.items():
        if rng is None:
            cell2 = ws.cell(row=row, column=c, value="")
        else:
            if func == "QUARTILE":
                q = 3 if "75" in label else 1
                formula = f"=QUARTILE({rng},{q})"
            else:
                formula = f"={func}({rng})"
            cell2 = ws.cell(row=row, column=c, value=formula)
            cell2.number_format = fmt
        cell2.font = FONT_DATA
        cell2.alignment = CENTER
        cell2.fill = LIGHTGREY


def build(inputs_path, output_path):
    inputs_path = Path(inputs_path)
    inputs = json.loads(inputs_path.read_text(encoding="utf-8"))
    peers_csv_rel = inputs["data_sources"]["comps"]
    peers_csv = inputs_path.parent / peers_csv_rel
    if not peers_csv.exists():
        raise FileNotFoundError(f"Peer CSV not found at {peers_csv}")
    with open(peers_csv, encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    wb = build_workbook(inputs, rows)
    wb.save(output_path)
    print(f"Wrote {output_path}")


def build_workbook(inputs, rows):
    company = inputs["company"]
    ticker = inputs["ticker"]
    as_of = inputs["as_of_date"]

    if not rows:
        raise ValueError("No peer rows provided")

    target_row = next((r for r in rows if r["Company"].lower() == company.lower()), rows[0])
    tickers = " vs ".join(f"{r['Company']} ({r['Ticker']})" for r in rows)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{company} Comps"

    ws.cell(row=1, column=1, value=f"COMPARABLE COMPANY ANALYSIS - {company.upper()} vs PEERS")
    ws.cell(row=2, column=1, value=tickers)
    ws.cell(row=3, column=1, value=f"As of {as_of} | Revenue/MCap/EV in local currency | Multiples and margins currency-neutral")

    section_header(ws, 5, "OPERATING METRICS", 7)
    col_header(ws, 6, ["Company", "Revenue (LCY M)", "Gross Margin", "EBITDA Margin", "Op Margin", "Net Margin", "Currency"])
    op_first = 7
    for i, row in enumerate(rows):
        r = op_first + i
        name = row["Company"]
        source = f"Data as of {as_of}. See inputs.json data_sources for URLs."
        rev_m = float(row["Revenue_TTM_LCY"]) / 1_000_000
        gp_m = float(row["Gross_Profit_TTM_LCY"]) / 1_000_000
        ebitda_m = float(row["EBITDA_TTM_LCY"]) / 1_000_000
        write_input(ws, r, 1, name, source_note=source)
        ws.cell(row=r, column=1).alignment = LEFT
        write_input(ws, r, 2, rev_m, source_note=source, fmt="#,##0")
        write_input(ws, r, 3, gp_m / rev_m, source_note=source, fmt="0.0%")
        write_input(ws, r, 4, ebitda_m / rev_m, source_note=source, fmt="0.0%")
        write_input(ws, r, 5, float(row["Operating_Margin_pct"]) / 100, fmt="0.0%")
        write_input(ws, r, 6, float(row["Net_Margin_pct"]) / 100, fmt="0.0%")
        ccell = ws.cell(row=r, column=7, value=row["Currency"])
        ccell.font = FONT_DATA
        ccell.alignment = CENTER

    op_last = op_first + len(rows) - 1
    stats_start = op_last + 2
    margin_ranges = {c: f"{get_column_letter(c)}{op_first}:{get_column_letter(c)}{op_last}" for c in [3, 4, 5, 6]}
    for offset, label in enumerate(["Max", "75th Percentile", "Median", "25th Percentile", "Min"]):
        r = stats_start + offset
        stats_row(ws, r, label, margin_ranges, "0.0%")
        for c in [2, 7]:
            cell = ws.cell(row=r, column=c, value="")
            cell.fill = LIGHTGREY

    val_section_row = stats_start + 6
    section_header(ws, val_section_row, "VALUATION MULTIPLES", 9)
    col_header(ws, val_section_row + 1, ["Company", "Market Cap (LCY M)", "Enterprise Value (LCY M)", "EV / Sales", "EV / EBITDA", "P/E (TTM)", "P/E (Fwd)", "Beta (5Y)", "Currency"])
    val_first = val_section_row + 2
    for i, row in enumerate(rows):
        r = val_first + i
        mcap_m = float(row["Market_Cap_LCY"]) / 1_000_000
        ev_m = float(row["Enterprise_Value_LCY"]) / 1_000_000
        write_input(ws, r, 1, row["Company"])
        ws.cell(row=r, column=1).alignment = LEFT
        write_input(ws, r, 2, mcap_m, fmt="#,##0")
        write_input(ws, r, 3, ev_m, fmt="#,##0")
        op_rev_cell = f"B{op_first + i}"
        write_formula(ws, r, 4, f"=C{r}/{op_rev_cell}", fmt='0.00"x"')
        op_ebmgn_cell = f"D{op_first + i}"
        write_formula(ws, r, 5, f"=C{r}/({op_rev_cell}*{op_ebmgn_cell})", fmt='0.00"x"')
        write_input(ws, r, 6, float(row["PE_Trailing"]), fmt='0.00"x"')
        write_input(ws, r, 7, float(row["PE_Forward"]), fmt='0.00"x"')
        write_input(ws, r, 8, float(row["Beta_5Y"]), fmt="0.00")
        ccell = ws.cell(row=r, column=9, value=row["Currency"])
        ccell.font = FONT_DATA
        ccell.alignment = CENTER

    val_last = val_first + len(rows) - 1
    val_stats_start = val_last + 2
    multiple_ranges = {c: f"{get_column_letter(c)}{val_first}:{get_column_letter(c)}{val_last}" for c in [4, 5, 6, 7, 8]}
    for offset, label in enumerate(["Max", "75th Percentile", "Median", "25th Percentile", "Min"]):
        r = val_stats_start + offset
        for col_idx, rng in multiple_ranges.items():
            fmt = "0.00" if col_idx == 8 else '0.00"x"'
            stats_row(ws, r, label, {col_idx: rng}, fmt)
        for c in [2, 3, 9]:
            cell = ws.cell(row=r, column=c, value="")
            cell.fill = LIGHTGREY

    notes_start = val_stats_start + 7
    ws.cell(row=notes_start, column=1, value="NOTES & METHODOLOGY").font = FONT_HDR_BLACK
    notes = [
        f"Target company: {company} ({ticker}). Compared against {len(rows) - 1} listed peers.",
        f"As of {as_of}. Data source: per data_sources in inputs.json.",
        "Multiples and margins are currency-neutral so cross-currency stats are meaningful. Absolute size in local currency.",
        "Stats include all rows (target + peers). Mentally exclude target to benchmark against peers only.",
        "EBITDA derived: Revenue * EBITDA margin per SKILL.md cross-reference rule.",
        "Accounting standards may differ (IFRS vs US GAAP) — verify before drawing precise conclusions.",
        "Limitations: snapshot, no historical trend. Add LFL and segment data for v2.",
    ]
    for i, n in enumerate(notes):
        ws.cell(row=notes_start + 1 + i, column=1, value=n).font = FONT_DATA
        ws.merge_cells(start_row=notes_start + 1 + i, start_column=1, end_row=notes_start + 1 + i, end_column=9)

    for c in range(1, 10):
        ws.column_dimensions[get_column_letter(c)].width = 22
    ws.column_dimensions["A"].width = 28

    wb.save(output_path)
    print(f"Wrote {output_path}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--inputs", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()
    build(args.inputs, args.output)


if __name__ == "__main__":
    main()
